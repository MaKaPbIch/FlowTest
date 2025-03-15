from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.db.models import Count, Avg, Q, F, FloatField, Max
from django.db.models.functions import TruncDate
from django.utils import timezone
from datetime import timedelta, datetime
from rest_framework import viewsets, status
from rest_framework.decorators import api_view, permission_classes, action
from rest_framework.permissions import IsAuthenticated, IsAdminUser
from rest_framework.response import Response
from rest_framework import serializers
from rest_framework.parsers import JSONParser, MultiPartParser, FormParser
from rest_framework.renderers import JSONRenderer
from rest_framework.negotiation import DefaultContentNegotiation
from rest_framework.views import APIView
from rest_framework_simplejwt.authentication import JWTAuthentication
from rest_framework.authentication import SessionAuthentication
from rest_framework.exceptions import ValidationError
from channels.layers import get_channel_layer
from asgiref.sync import sync_to_async, async_to_sync
import json
import logging
import os
import io
import csv
import asyncio
from .models import (
    Project, Folder, TestCase, Role, CustomUser, Permission,
    AutomationProject, TestRun, SchedulerEvent, ReportTemplate,
    TestReport
)
from .serializers import (
    ProjectSerializer, FolderSerializer, TestCaseSerializer,
    RoleSerializer, CustomUserSerializer, PermissionSerializer, AutomationProjectSerializer,
    TestRunSerializer, SchedulerEventSerializer, ReportTemplateSerializer,
    ReportMetricsSerializer, ReportChartDataSerializer,
    TestReportSerializer, AnalyticsResponseSerializer
)
from .services.automation_service import AutomationService
from .services.repository_service import RepositoryService
from .services.scheduler_service import SchedulerService
from .tasks import execute_test
from FlowTest.celery import app

# Валидаторы для аватара
def validate_avatar_size(value):
    if value and value.size > 2 * 1024 * 1024:  # 2MB
        raise ValidationError("Avatar size cannot exceed 2MB")

def validate_avatar_extension(value):
    if value:
        valid_extensions = ['.jpg', '.jpeg', '.png', '.gif']
        ext = os.path.splitext(value.name)[1].lower()
        if ext not in valid_extensions:
            raise ValidationError(f"Unsupported file extension. Allowed extensions are: {', '.join(valid_extensions)}")

# Сериализатор пользователя
class CustomUserSerializer(serializers.ModelSerializer):
    avatar = serializers.SerializerMethodField()
    password = serializers.CharField(write_only=True, required=False)

    def get_avatar(self, obj):
        if obj.avatar:
            request = self.context.get('request')
            if request:
                return request.build_absolute_uri(obj.avatar.url)
            return obj.avatar.url
        return None

    class Meta:
        model = CustomUser
        fields = ['id', 'username', 'password', 'email', 'first_name', 'last_name', 
                  'middle_name', 'language', 'phone_number', 'avatar', 'theme', 'role']
        extra_kwargs = {
            'password': {'write_only': True}
        }

    def update(self, instance, validated_data):
        if 'avatar' in validated_data:
            if instance.avatar:
                instance.avatar.delete(save=False)
        return super().update(instance, validated_data)

# Создание тестового прогона
def create_test_run(test_case):
    test_run = TestRun(
        test_case=test_case,
        status='pending',
        started_at=timezone.now(),
        log_output='Test execution started...\n'
    )
    test_run.save()
    return test_run

@sync_to_async
def get_test_run(test_run_id):
    test_run = TestRun.objects.get(id=test_run_id)
    test_run.test_case
    return test_run

@sync_to_async
def save_test_run(test_run):
    test_run.save()

# Асинхронный запуск теста
async def run_test_in_thread(test_run_id):
    try:
        print(f"Starting test execution in thread for test run {test_run_id}")
        test_run = await sync_to_async(TestRun.objects.get)(id=test_run_id)
        test_case = test_run.test_case
        
        test_run.status = 'running'
        test_run.log_output = 'Initializing browser...\n'
        await sync_to_async(test_run.save)()
        
        from playwright.sync_api import sync_playwright
        import threading
        
        def run_playwright_test():
            try:
                with sync_playwright() as p:
                    browser = p.chromium.launch(headless=False, args=['--start-maximized'])
                    try:
                        context = browser.new_context(viewport={'width': 1920, 'height': 1080})
                        page = context.new_page()
                        test_locals = {
                            'page': page,
                            'context': context,
                            'browser': browser,
                            'test_run': test_run,
                            'logger': None,
                        }
                        exec(test_case.test_code, test_locals)
                        test_run.status = 'success'
                        test_run.finished_at = timezone.now()
                        test_run.duration = (test_run.finished_at - test_run.started_at).total_seconds()
                        test_run.log_output += '\nTest completed successfully\n'
                    except Exception as e:
                        import traceback
                        error_details = f'\nTest failed: {str(e)}\n{traceback.format_exc()}'
                        test_run.status = 'failure'
                        test_run.finished_at = timezone.now()
                        test_run.duration = (test_run.finished_at - test_run.started_at).total_seconds()
                        test_run.error_message = str(e)
                        test_run.log_output += error_details
                    finally:
                        if 'context' in locals():
                            context.close()
                        browser.close()
            except Exception as e:
                print(f"Error initializing Playwright: {str(e)}", exc_info=True)
                test_run.status = 'error'
                test_run.finished_at = timezone.now()
                test_run.duration = (test_run.finished_at - test_run.started_at).total_seconds()
                test_run.error_message = f'Failed to initialize Playwright: {str(e)}'
                test_run.log_output += '\nFailed to initialize browser\n'
        
        thread = threading.Thread(target=run_playwright_test)
        thread.start()
        thread.join()
        await sync_to_async(test_run.save)()
    except Exception as e:
        print(f"Error during test execution: {str(e)}", exc_info=True)
        try:
            test_run = await sync_to_async(TestRun.objects.get)(id=test_run_id)
            test_run.status = 'error'
            test_run.finished_at = timezone.now()
            test_run.duration = (test_run.finished_at - test_run.started_at).total_seconds()
            test_run.error_message = f'Failed to initialize test: {str(e)}'
            test_run.log_output = 'Test initialization failed\n'
            await sync_to_async(test_run.save)()
        except Exception as e2:
            print(f"Failed to update test run status: {str(e2)}", exc_info=True)

def run_test_in_thread_sync(test_run_id):
    asyncio.run(run_test_in_thread(test_run_id))

# ViewSet для TestCase
class TestCaseViewSet(viewsets.ModelViewSet):
    queryset = TestCase.objects.all()
    serializer_class = TestCaseSerializer
    permission_classes = [IsAuthenticated]
    authentication_classes = [JWTAuthentication, SessionAuthentication]

    def get_queryset(self):
        queryset = TestCase.objects.all()
        folder_id = self.request.query_params.get('folder', None)
        project_id = self.request.query_params.get('project', None)
        if folder_id is not None:
            queryset = queryset.filter(folder_id=folder_id)
        if project_id is not None:
            queryset = queryset.filter(project_id=project_id)
        return queryset

    def perform_create(self, serializer):
        serializer.save(author=self.request.user)

    @action(detail=True, methods=['post'])
    def execute(self, request, pk=None):
        return Response({
            'status': 'error',
            'error': 'Please use /api/execute-test/<test_id>/ endpoint'
        }, status=400)

    @action(detail=True, methods=['get'])
    def get_test_status(self, request, pk=None):
        print(f"Getting status for test run {pk}")
        try:
            test_run = TestRun.objects.get(id=pk)
            return Response({
                'status': test_run.status,
                'started_at': test_run.started_at,
                'finished_at': test_run.finished_at,
                'duration': test_run.duration,
                'output': test_run.log_output,
                'error': test_run.error_message
            })
        except TestRun.DoesNotExist:
            return Response({'status': 'error', 'error': f'Test run {pk} not found'}, status=404)
        except Exception as e:
            print(f"Error getting test status: {str(e)}", exc_info=True)
            return Response({'status': 'error', 'error': str(e)}, status=500)

    @action(detail=True, methods=['post'])
    def add_tag(self, request, pk=None):
        test_case = self.get_object()
        tag = request.data.get('tag')
        if not tag:
            return Response({'error': 'Tag is required'}, status=400)
        tags = test_case.tags or []
        if tag not in tags:
            tags.append(tag)
            test_case.tags = tags
            test_case.save()
        return Response({'tags': test_case.tags})

    @action(detail=True, methods=['post'])
    def remove_tag(self, request, pk=None):
        test_case = self.get_object()
        tag = request.data.get('tag')
        if not tag:
            return Response({'error': 'Tag is required'}, status=400)
        tags = test_case.tags or []
        if tag in tags:
            tags.remove(tag)
            test_case.tags = tags
            test_case.save()
        return Response({'tags': test_case.tags})

    @action(detail=True, methods=['post'])
    def update_steps(self, request, pk=None):
        test_case = self.get_object()
        steps = request.data.get('steps')
        if steps is None:
            return Response({'error': 'Steps are required'}, status=400)
        test_case.steps = steps
        test_case.save()
        return Response({'steps': test_case.steps})

    @action(detail=True, methods=['post'])
    def add_step(self, request, pk=None):
        test_case = self.get_object()
        step = request.data.get('step')
        if not step:
            return Response({'error': 'Step is required'}, status=400)
        steps = test_case.steps or []
        steps.append(step)
        test_case.steps = steps
        test_case.save()
        return Response({'steps': test_case.steps})

    @action(detail=True, methods=['post'])
    def remove_step(self, request, pk=None):
        test_case = self.get_object()
        step_index = request.data.get('step_index')
        if step_index is None:
            return Response({'error': 'Step index is required'}, status=400)
        steps = test_case.steps or []
        if 0 <= step_index < len(steps):
            steps.pop(step_index)
            test_case.steps = steps
            test_case.save()
            return Response({'steps': test_case.steps})
        return Response({'error': 'Invalid step index'}, status=400)

    @action(detail=True, methods=['post'])
    def update_test_code(self, request, pk=None):
        test_case = self.get_object()
        test_code = request.data.get('test_code')
        if test_code is None:
            return Response({'error': 'Test code is required'}, status=400)
        test_case.test_code = test_code
        test_case.save()
        return Response({'test_code': test_case.test_code})

    @action(detail=False, methods=['GET'])
    def analytics(self, request):
        project_id = request.query_params.get('project_id')
        date_from = request.query_params.get('from')
        date_to = request.query_params.get('to')
        
        if not project_id:
            return Response({'error': 'project_id is required'}, status=400)
            
        queryset = self.filter_queryset(self.get_queryset())
        queryset = queryset.filter(project_id=project_id)
        
        if date_from and date_to:
            queryset = queryset.filter(created_at__date__gte=date_from, created_at__date__lte=date_to)
            
        test_runs = TestRun.objects.filter(test_case__in=queryset)
        if date_from and date_to:
            test_runs = test_runs.filter(started_at__date__gte=date_from, started_at__date__lte=date_to)
        stats = {
            'total_cases': queryset.count(),
            'status_distribution': {
                'passed': test_runs.filter(status='success').count(),
                'failed': test_runs.filter(status='failure').count(),
                'error': test_runs.filter(status='error').count(),
                'not_run': queryset.exclude(id__in=test_runs.values('test_case')).count()
            },
            'user_activity': list(
                queryset.values('author__username')
                .annotate(created=Count('id'), updated=Count('id', filter=Q(updated_at__gt=F('created_at'))))
                .exclude(author__username=None)
                .values('author__username', 'created', 'updated')
                .order_by('-created')
            ),
            'daily_trends': list(
                test_runs.annotate(date=TruncDate('started_at'))
                .values('date')
                .annotate(total=Count('id'), passed=Count('id', filter=Q(status='success')),
                          failed=Count('id', filter=Q(status='failure')), error=Count('id', filter=Q(status='error')))
                .order_by('date')
            )
        }
        return Response(stats)

    @action(detail=False, methods=['POST'])
    def export_analytics(self, request):
        try:
            format = request.data.get('format', 'csv')
            data_types = request.data.get('data', [])
            date_from = request.data.get('date_from')
            date_to = request.data.get('date_to')
            analytics_data = self._get_analytics_data(date_from, date_to)
            export_data = {data_type: analytics_data[data_type] for data_type in data_types if data_type in analytics_data}
            if format == 'csv':
                return self._export_to_csv(export_data)
            elif format == 'excel':
                return self._export_to_excel(export_data)
            elif format == 'pdf':
                return self._export_to_pdf(export_data)
            return Response({'error': 'Unsupported format'}, status=400)
        except Exception as e:
            print(f"Error in export_analytics: {str(e)}", exc_info=True)
            return Response({'error': str(e)}, status=500)

    def _get_analytics_data(self, date_from=None, date_to=None):
        queryset = self.filter_queryset(self.get_queryset())
        if date_from and date_to:
            queryset = queryset.filter(created_at__date__gte=date_from, created_at__date__lte=date_to)
        test_runs = TestRun.objects.filter(test_case__in=queryset)
        if date_from and date_to:
            test_runs = test_runs.filter(started_at__date__gte=date_from, started_at__date__lte=date_to)
        return {
            'test_status': {
                'passed': test_runs.filter(status='success').count(),
                'failed': test_runs.filter(status='failure').count(),
                'error': test_runs.filter(status='error').count(),
                'not_run': queryset.exclude(id__in=test_runs.values('test_case')).count()
            },
            'user_activity': list(
                queryset.values('author__username')
                .annotate(created=Count('id'), updated=Count('id', filter=Q(updated_at__gt=F('created_at'))))
                .exclude(author__username=None)
                .values('author__username', 'created', 'updated')
                .order_by('-created')
            ),
            'daily_trends': list(
                test_runs.annotate(date=TruncDate('started_at'))
                .values('date')
                .annotate(total=Count('id'), passed=Count('id', filter=Q(status='success')),
                          failed=Count('id', filter=Q(status='failure')), error=Count('id', filter=Q(status='error')))
                .order_by('date')
            )
        }

    def _export_to_csv(self, data):
        output = io.StringIO()
        writer = csv.writer(output)
        for data_type, values in data.items():
            writer.writerow([f'=== {data_type.upper()} ==='])
            if isinstance(values, dict):
                for key, value in values.items():
                    writer.writerow([key, value])
            elif isinstance(values, list) and values:
                writer.writerow(values[0].keys())
                for item in values:
                    writer.writerow(item.values())
            writer.writerow([])
        response = HttpResponse(output.getvalue(), content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="analytics_export.csv"'
        return response

    def _export_to_excel(self, data):
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        wb = openpyxl.Workbook()
        for data_type, values in data.items():
            ws = wb.create_sheet(title=data_type.replace('_', ' ').title())
            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
            if isinstance(values, dict):
                ws['A1'] = 'Metric'
                ws['B1'] = 'Value'
                ws['A1'].font = header_font
                ws['B1'].font = header_font
                ws['A1'].fill = header_fill
                ws['B1'].fill = header_fill
                row = 2
                for key, value in values.items():
                    ws[f'A{row}'] = key
                    ws[f'B{row}'] = value
                    row += 1
            elif isinstance(values, list) and values:
                headers = list(values[0].keys())
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                for row, item in enumerate(values, 2):
                    for col, value in enumerate(item.values(), 1):
                        ws.cell(row=row, column=col, value=value)
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="analytics_export.xlsx"'
        wb.save(response)
        return response

    def _export_to_pdf(self, data):
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()
        for data_type, values in data.items():
            elements.append(Paragraph(data_type.replace('_', ' ').title(), styles['Heading1']))
            elements.append(Spacer(1, 12))
            if isinstance(values, dict):
                table_data = [[Paragraph('Metric', styles['Heading2']), Paragraph('Value', styles['Heading2'])]]
                for key, value in values.items():
                    table_data.append([key, str(value)])
            elif isinstance(values, list) and values:
                headers = list(values[0].keys())
                table_data = [[Paragraph(h.replace('_', ' ').title(), styles['Heading2']) for h in headers]]
                for item in values:
                    table_data.append([str(v) for v in item.values()])
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F81BD')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 12),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(table)
            elements.append(Spacer(1, 20))
        doc.build(elements)
        pdf = buffer.getvalue()
        buffer.close()
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="analytics_export.pdf"'
        response.write(pdf)
        return response

# ViewSet для TestRun
class TestRunViewSet(viewsets.ModelViewSet):
    queryset = TestRun.objects.all()
    serializer_class = TestRunSerializer
    permission_classes = [IsAuthenticated]
    authentication_classes = [JWTAuthentication, SessionAuthentication]

    def get_queryset(self):
        queryset = TestRun.objects.all()
        test_case_id = self.request.query_params.get('test_case', None)
        if test_case_id is not None:
            queryset = queryset.filter(test_case_id=test_case_id)
        return queryset

    @action(detail=True, methods=['get'], url_path='status')
    def get_test_status(self, request, pk=None):
        print(f"Getting status for test run {pk}")
        try:
            test_run = TestRun.objects.get(id=pk)
            return Response({
                'status': test_run.status,
                'started_at': test_run.created_at,
                'finished_at': test_run.finished_at,
                'duration': test_run.duration,
                'output': test_run.log_output,
                'error': test_run.error_message
            })
        except TestRun.DoesNotExist:
            return Response({'status': 'error', 'error': f'Test run {pk} not found'}, status=404)
        except Exception as e:
            print(f"Error getting test status: {str(e)}", exc_info=True)
            return Response({'status': 'error', 'error': str(e)}, status=500)

# ViewSet для SchedulerEvent
class SchedulerEventViewSet(viewsets.ModelViewSet):
    queryset = SchedulerEvent.objects.all()
    serializer_class = SchedulerEventSerializer
    permission_classes = [IsAuthenticated]
    authentication_classes = [JWTAuthentication, SessionAuthentication]

    def delete(self, request, pk=None):
        event = self.get_object()
        event.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=True, methods=['post'])
    def create_recurrent_event(self, request, pk=None):
        event = self.get_object()
        recurrence_type = request.data.get('recurrence_type')
        recurrence_count = request.data.get('recurrence_count', 1)
        if not recurrence_type or recurrence_count < 1:
            return Response({'error': 'Invalid recurrence parameters'}, status=status.HTTP_400_BAD_REQUEST)
        created_events = []
        current_time = event.scheduled_time
        for i in range(recurrence_count):
            if recurrence_type == 'daily':
                next_time = current_time + timedelta(days=1)
            elif recurrence_type == 'weekly':
                next_time = current_time + timedelta(weeks=1)
            elif recurrence_type == 'monthly':
                next_time = current_time + timedelta(days=30)
            else:
                return Response({'error': 'Invalid recurrence type'}, status=status.HTTP_400_BAD_REQUEST)
            new_event = SchedulerEvent.objects.create(project=event.project, scheduled_time=next_time, status='pending')
            created_events.append(new_event)
            current_time = next_time
        serializer = SchedulerEventSerializer(created_events, many=True)
        return Response(serializer.data)

# ViewSet для Role
class PermissionViewSet(viewsets.ModelViewSet):
    queryset = Permission.objects.all()
    serializer_class = PermissionSerializer
    permission_classes = [IsAuthenticated]
    authentication_classes = [JWTAuthentication, SessionAuthentication]
    
    def get_queryset(self):
        queryset = Permission.objects.all()
        
        # Filter by category if provided
        category = self.request.query_params.get('category', None)
        if category:
            queryset = queryset.filter(category=category)
            
        return queryset

class RoleViewSet(viewsets.ModelViewSet):
    queryset = Role.objects.all()
    serializer_class = RoleSerializer
    permission_classes = [IsAuthenticated]
    authentication_classes = [JWTAuthentication, SessionAuthentication]

# ViewSet для CustomUser
class CustomUserViewSet(viewsets.ModelViewSet):
    queryset = CustomUser.objects.all()
    serializer_class = CustomUserSerializer
    permission_classes = [IsAuthenticated]
    parser_classes = [JSONParser, MultiPartParser, FormParser]
    renderer_classes = [JSONRenderer]
    authentication_classes = [JWTAuthentication, SessionAuthentication]
    
    def get_permissions(self):
        """
        Override to set custom permissions per action
        """
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            # Only staff/admin users can create, update or delete users
            return [IsAuthenticated(), IsAdminUser()]
        return super().get_permissions()

    def get_queryset(self):
        if not self.request.user.is_authenticated:
            return CustomUser.objects.none()
        
        # Allow admins to see all users
        if self.request.user.is_staff or self.request.user.is_superuser:
            return CustomUser.objects.all()
            
        # Regular users can only see themselves
        return CustomUser.objects.filter(id=self.request.user.id)

    @action(detail=False, methods=['get'])
    def get_current_user(self, request):
        if not request.user.is_authenticated:
            return Response({"error": "Authentication required"}, status=status.HTTP_401_UNAUTHORIZED)
        
        user = request.user
        print(f"Avatar path: {user.avatar.path if user.avatar else 'None'}")
        print(f"Avatar URL: {user.avatar.url if user.avatar else 'None'}")
        print(f"Avatar name: {user.avatar.name if user.avatar else 'None'}")
        
        serializer = self.get_serializer(user, context={'request': request})
        data = serializer.data
        print(f"Serialized data: {data}")
        return Response(data)

    @action(detail=True, methods=['patch'])
    def update_avatar(self, request, pk=None):
        try:
            if str(request.user.id) != str(pk):
                return Response({"error": "You can only update your own avatar"}, status=status.HTTP_403_FORBIDDEN)
            user = self.get_object()
            if 'avatar' not in request.FILES:
                return Response({"error": "No avatar file provided"}, status=status.HTTP_400_BAD_REQUEST)
            avatar_file = request.FILES['avatar']
            validate_avatar_size(avatar_file)
            validate_avatar_extension(avatar_file)
            if user.avatar:
                user.avatar.delete(save=False)
            user.avatar = avatar_file
            user.save()
            serializer = self.get_serializer(user, context={'request': request})
            return Response(serializer.data)
        except ValidationError as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def update(self, request, *args, **kwargs):
        try:
            partial = kwargs.pop('partial', False)
            instance = self.get_object()
            if instance.id != request.user.id:
                return Response({"error": "You can only update your own profile"}, status=status.HTTP_403_FORBIDDEN)
            serializer = self.get_serializer(instance, data=request.data, partial=partial, context={'request': request})
            serializer.is_valid(raise_exception=True)
            self.perform_update(serializer)
            return Response(serializer.data)
        except ValidationError as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# ViewSet для AutomationProject
class AutomationProjectViewSet(viewsets.ModelViewSet):
    queryset = AutomationProject.objects.all()
    serializer_class = AutomationProjectSerializer
    permission_classes = [IsAuthenticated]
    parser_classes = [JSONParser]
    http_method_names = ['get', 'post', 'put', 'patch', 'delete']
    authentication_classes = [JWTAuthentication, SessionAuthentication]

    def create(self, request, *args, **kwargs):
        print("Request data:", request.data)
        print("Request headers:", request.headers)
        try:
            return super().create(request, *args, **kwargs)
        except Exception as e:
            print("Error in create:", str(e))
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    def perform_create(self, serializer):
        print("Validated data:", serializer.validated_data)
        if not serializer.validated_data.get('project'):
            raise ValidationError("Project is required")
        serializer.save()

    @action(detail=True, methods=['post'])
    def sync(self, request, pk=None):
        project = self.get_object()
        automation_service = AutomationService()
        try:
            automation_service.sync_repository(project)
            return Response({'status': 'success', 'message': 'Repository synchronized'})
        except Exception as e:
            return Response({'status': 'error', 'message': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['get'])
    def tests(self, request, pk=None):
        project = self.get_object()
        tests = project.tests.all()
        serializer = TestCaseSerializer(tests, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def run_tests(self, request, pk=None):
        project = self.get_object()
        test_ids = request.data.get('test_ids', [])
        if not test_ids:
            return Response({'error': 'No tests selected'}, status=status.HTTP_400_BAD_REQUEST)
        automation_service = AutomationService()
        try:
            results = automation_service.run_tests(project, test_ids)
            return Response({'status': 'success', 'results': results})
        except Exception as e:
            return Response({'status': 'error', 'message': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['post'])
    def run_all(self, request, pk=None):
        project = self.get_object()
        automation_service = AutomationService()
        try:
            results = automation_service.run_all_tests(project)
            return Response({'status': 'success', 'results': results})
        except Exception as e:
            return Response({'status': 'error', 'message': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['post'])
    def schedule(self, request, pk=None):
        project = self.get_object()
        schedule_data = request.data
        serializer = SchedulerEventSerializer(data=schedule_data)
        if serializer.is_valid():
            serializer.save(project=project)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

# ViewSet для Project
class ProjectViewSet(viewsets.ModelViewSet):
    queryset = Project.objects.all()
    serializer_class = ProjectSerializer
    permission_classes = [IsAuthenticated]
    authentication_classes = [JWTAuthentication, SessionAuthentication]

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['get'])
    def folders(self, request, pk=None):
        project = self.get_object()
        folders = project.folders.all()
        serializer = FolderSerializer(folders, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['get'])
    def folders_and_test_cases(self, request, pk=None):
        project = self.get_object()
        folders = project.folders.all()
        folder_data = []
        test_cases = project.test_cases.all()
        test_cases_by_folder = {}
        for test_case in test_cases:
            folder_id = test_case.folder_id if test_case.folder else None
            if folder_id not in test_cases_by_folder:
                test_cases_by_folder[folder_id] = []
            test_cases_by_folder[folder_id].append(test_case)
        for folder in folders:
            folder_serializer = FolderSerializer(folder)
            folder_test_cases = test_cases_by_folder.get(folder.id, [])
            test_case_serializer = TestCaseSerializer(folder_test_cases, many=True)
            folder_info = folder_serializer.data
            folder_info['test_cases'] = test_case_serializer.data
            folder_data.append(folder_info)
        unassigned_test_cases = test_cases_by_folder.get(None, [])
        if unassigned_test_cases:
            folder_data.append({
                'id': None,
                'name': 'Unassigned',
                'description': 'Test cases without folder',
                'project': project.id,
                'test_cases': TestCaseSerializer(unassigned_test_cases, many=True).data
            })
        return Response(folder_data)

# ViewSet для Folder
class FolderViewSet(viewsets.ModelViewSet):
    queryset = Folder.objects.all()
    serializer_class = FolderSerializer
    permission_classes = [IsAuthenticated]
    authentication_classes = [JWTAuthentication, SessionAuthentication]

    def get_queryset(self):
        queryset = Folder.objects.all()
        project_id = self.request.query_params.get('project', None)
        if project_id is not None:
            queryset = queryset.filter(project_id=project_id)
        return queryset

    @action(detail=True, methods=['get'])
    def test_cases(self, request, pk=None):
        folder = self.get_object()
        test_cases = folder.test_cases.all()
        serializer = TestCaseSerializer(test_cases, many=True)
        return Response(serializer.data)

# Миксин для асинхронной аутентификации
class AsyncAuthMixin(APIView):
    renderer_classes = [JSONRenderer]
    content_negotiation_class = DefaultContentNegotiation

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.headers = {}

    def prepare_request(self, request):
        request.query_params = request.GET
        request.data = getattr(request, '_body', request.POST)
        request._request = request
        request._full_data = request.data
        return request

    async def dispatch(self, request, *args, **kwargs):
        self.args = args
        self.kwargs = kwargs
        self.request = self.prepare_request(request)
        self.headers = {}
        self.format_kwarg = None
        if not hasattr(request, '_user'):
            request._user = await async_to_sync(self._authenticate_credentials)(request)
            request.user = request._user
        request.accepted_renderer = self.renderer_classes[0]()
        request.accepted_media_type = request.accepted_renderer.media_type
        try:
            handler = getattr(self, request.method.lower())
            response = await handler(request, *args, **kwargs)
            if not isinstance(response, Response):
                response = Response(response)
            response.accepted_renderer = request.accepted_renderer
            response.accepted_media_type = request.accepted_media_type
            response.renderer_context = {
                'view': self,
                'args': self.args,
                'kwargs': self.kwargs,
                'request': request
            }
            return response
        except Exception as exc:
            response = self.handle_exception(exc)
            response.accepted_renderer = request.accepted_renderer
            response.accepted_media_type = request.accepted_media_type
            response.renderer_context = {
                'view': self,
                'args': self.args,
                'kwargs': self.kwargs,
                'request': request
            }
            return response

    def _authenticate_credentials(self, request):
        for authenticator in self.authentication_classes:
            try:
                user_auth_tuple = authenticator().authenticate(request)
                if user_auth_tuple is not None:
                    return user_auth_tuple[0]
            except:
                continue
        return None

# Асинхронное выполнение теста
class TestExecutionView(AsyncAuthMixin):
    permission_classes = [IsAuthenticated]
    authentication_classes = [JWTAuthentication, SessionAuthentication]

    async def post(self, request, test_id):
        print(f"Executing test {test_id}")
        try:
            test_case = await sync_to_async(TestCase.objects.get)(id=test_id)
            print(f"Found test case: {test_case}")
            test_run = await sync_to_async(TestRun.objects.create)(
                test_case=test_case,
                status='pending',
                started_at=timezone.now()
            )
            print(f"Created test run: {test_run}")
            from .tasks import execute_test
            task = execute_test.delay(test_run.id)
            print(f"Started Celery task: {task.id}")
            return Response({
                'status': 'success',
                'message': 'Test execution started',
                'test_run_id': test_run.id,
                'task_id': task.id
            })
        except TestCase.DoesNotExist:
            print(f"Test case {test_id} not found")
            return Response({'status': 'error', 'error': f'Test case {test_id} not found'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            print(f"Error executing test {test_id}: {str(e)}", exc_info=True)
            return Response({'status': 'error', 'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    async def options(self, request, *args, **kwargs):
        response = Response()
        response["Access-Control-Allow-Origin"] = "http://127.0.0.1:8080"
        response["Access-Control-Allow-Methods"] = "POST, OPTIONS"
        response["Access-Control-Allow-Headers"] = "Content-Type, Authorization"
        return response

# Асинхронный статус теста
class TestStatusView(AsyncAuthMixin):
    permission_classes = [IsAuthenticated]
    authentication_classes = [JWTAuthentication, SessionAuthentication]

    async def get(self, request, test_run_id):
        print(f"Getting status for test run {test_run_id}")
        try:
            test_run = await sync_to_async(TestRun.objects.select_related('test_case').get)(id=test_run_id)
            response_data = {
                'status': test_run.status,
                'test_case_id': test_run.test_case.id if test_run.test_case else None,
                'test_case_title': test_run.test_case.title if test_run.test_case else None,
                'started_at': test_run.started_at.isoformat() if test_run.started_at else None,
                'finished_at': test_run.finished_at.isoformat() if test_run.finished_at else None,
                'duration': test_run.duration,
                'error_message': test_run.error_message,
                'log_output': test_run.log_output
            }
            return Response(response_data)
        except TestRun.DoesNotExist:
            return Response({'status': 'error', 'error': f'Test run {test_run_id} not found'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            print(f"Error getting test status: {str(e)}", exc_info=True)
            return Response({'status': 'error', 'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    async def options(self, request, *args, **kwargs):
        response = Response()
        response["Access-Control-Allow-Origin"] = "http://127.0.0.1:8080"
        response["Access-Control-Allow-Methods"] = "GET, OPTIONS"
        response["Access-Control-Allow-Headers"] = "Content-Type, Authorization"
        return response

# Проверка существования теста
class CheckTestExistenceView(APIView):
    permission_classes = [IsAuthenticated]
    authentication_classes = [JWTAuthentication, SessionAuthentication]

    def get(self, request, test_id):
        try:
            result = RepositoryService.check_test_existence(test_id)
            return Response(result)
        except Exception as e:
            print(f"Error checking test existence: {str(e)}", exc_info=True)
            return Response({'status': 'error', 'error': str(e)}, status=500)

# API для запуска теста
@api_view(['POST'])
def execute_test_api(request, test_id):
    try:
        test_case = TestCase.objects.get(id=test_id)
        test_run = TestRun.objects.create(test_case=test_case, status='pending')
        return Response({'test_run_id': test_run.id})
    except TestCase.DoesNotExist:
        return Response({'error': 'Test not found'}, status=404)
    except Exception as e:
        return Response({'error': str(e)}, status=500)

# Проверка существования теста в репозитории
@api_view(['GET'])
def check_test_existence(request, test_id):
    result = RepositoryService.check_test_existence(test_id)
    return Response(result)

# Распределение результатов тестов
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def results_distribution(request, project_id=None):
    try:
        project_id = request.GET.get('project_id')
        
        queryset = TestRun.objects.all()
        if project_id:
            queryset = queryset.filter(test_case__folder__project_id=project_id)
            
        stats = queryset.values('status').annotate(
            count=Count('id')
        ).order_by('-count')
        
        return Response(list(stats))
    except Exception as e:
        return Response({'error': str(e)}, status=400)

# Распределение тестов по приоритетам
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def priority_distribution(request):
    try:
        project_id = request.GET.get('project_id')
        
        queryset = TestCase.objects.all()
        if project_id:
            queryset = TestCase.objects.filter(folder__project_id=project_id)
            
        stats = queryset.values('priority').annotate(
            count=Count('id')
        ).order_by('-count')
        
        return Response(list(stats))
    except Exception as e:
        return Response({'error': str(e)}, status=400)

# Нестабильность тестов
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def test_flakiness(request):
    try:
        project_id = request.query_params.get('project_id')
        if not project_id:
            return Response({'error': 'project_id is required'}, status=400)

        # Get test cases for the project
        queryset = TestCase.objects.filter(folder__project_id=project_id)

        # Find tests with alternating success/failure patterns
        flaky_tests = []
        for test in queryset:
            runs = TestRun.objects.filter(test_case=test).order_by('-started_at')[:10]
            if runs.count() >= 5:  # Minimum 5 runs for analysis
                statuses = list(runs.values_list('status', flat=True))
                changes = sum(1 for i in range(len(statuses)-1) if statuses[i] != statuses[i+1])
                if changes >= 2:  # Minimum 2 status changes
                    flaky_tests.append({
                        'test_id': test.id,
                        'title': test.title,
                        'changes': changes,
                        'last_runs': statuses
                    })

        return Response(sorted(flaky_tests, key=lambda x: x['changes'], reverse=True))
    except Exception as e:
        return Response({'error': str(e)}, status=500)

# Статистика выполнения тестов
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def test_execution_stats(request, project_id=None):
    try:
        if not project_id:
            project_id = request.query_params.get('project_id')
            if not project_id:
                return Response({
                    'total_tests': 0,
                    'total_executions': 0,
                    'status_distribution': {
                        'passed': 0,
                        'failed': 0,
                        'error': 0,
                        'running': 0,
                        'pending': 0
                    }
                })

        # Count test cases
        total_tests = TestCase.objects.filter(folder__project_id=project_id).count()

        # Get all test runs
        test_runs = TestRun.objects.filter(test_case__folder__project_id=project_id)
        total_executions = test_runs.count()

        # Calculate distribution by status
        status_distribution = {}
        status_counts = test_runs.values('status').annotate(count=Count('id'))
        for item in status_counts:
            status_distribution[item['status']] = item['count']

        # Ensure all statuses have a value
        for status in ['passed', 'failed', 'error', 'skipped', 'running', 'pending']:
            if status not in status_distribution:
                status_distribution[status] = 0

        # Calculate average execution time (in seconds)
        avg_time = test_runs.aggregate(avg_time=Avg(
            ExpressionWrapper(
                F('finished_at') - F('started_at'),
                output_field=DurationField()
            )
        ))['avg_time']

        avg_execution_time = 0
        if avg_time:
            avg_execution_time = avg_time.total_seconds()

        stats = {
            'total_tests': total_tests,
            'total_executions': total_executions,
            'status_distribution': status_distribution,
            'avg_execution_time': avg_execution_time
        }

        return Response(stats)
    except Exception as e:
        return Response({
            'total_tests': 0,
            'total_executions': 0,
            'status_distribution': {
                'passed': 0,
                'failed': 0,
                'error': 0,
                'running': 0,
                'pending': 0
            },
            'avg_execution_time': 0
        })

# Топ Контрибьюторов
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def top_contributors(request):
    try:
        project_id = request.query_params.get('project_id')
        if not project_id:
            return Response([])
        
        queryset = TestCase.objects.all()
        queryset = queryset.filter(folder__project_id=project_id)
            
        # Получаем топ-авторов по количеству созданных тест-кейсов
        authors = queryset.values('created_by__username').annotate(
            count=Count('id')
        ).order_by('-count')[:10]
        
        # Преобразуем в нужный формат
        result = []
        for author in authors:
            username = author['created_by__username']
            result.append({
                'author': username if username else 'Unknown',
                'count': author['count']
            })
        
        return Response(result)
    except Exception as e:
        # Логируем ошибку, но возвращаем пустой список, чтобы не ломать интерфейс
        print(f"Error in top_contributors: {str(e)}")
        return Response([])

# ViewSet для ReportTemplate
class ReportTemplateViewSet(viewsets.ModelViewSet):
    queryset = ReportTemplate.objects.all()
    serializer_class = ReportTemplateSerializer
    permission_classes = [IsAuthenticated]
    authentication_classes = [JWTAuthentication, SessionAuthentication]
    
    def get_queryset(self):
        """
        Возвращает только неудаленные шаблоны
        """
        # Фильтруем, чтобы не показывать удаленные шаблоны
        queryset = ReportTemplate.objects.filter(is_deleted=False)
        
        # Фильтруем по проекту, если указан параметр project
        project_id = self.request.query_params.get('project')
        if project_id:
            queryset = queryset.filter(project_id=project_id)
            
        return queryset

    def perform_create(self, serializer):
        # Получаем project_id из данных запроса
        project_id = self.request.data.get('project')
        if not project_id:
            # Для обратной совместимости проверяем project_id
            project_id = self.request.data.get('project_id')
            
        serializer.save(created_by=self.request.user, project_id=project_id)

    def destroy(self, request, *args, **kwargs):
        try:
            instance = self.get_object()
            template_id = instance.id
            template_name = instance.name
            
            # Проверяем права доступа (только создатель или админ может удалить)
            if not request.user.is_superuser and instance.created_by != request.user:
                return Response(
                    {'error': 'You do not have permission to delete this template'}, 
                    status=status.HTTP_403_FORBIDDEN
                )
                
            # Логируем попытку удаления
            print(f"Marking template {template_id} '{template_name}' as deleted by {request.user.username}")
            
            # Вместо физического удаления, помечаем запись как удаленную
            instance.is_deleted = True
            instance.save()
            
            # Логируем успешное удаление
            print(f"Template {template_id} '{template_name}' was successfully marked as deleted")
            
            return Response(status=status.HTTP_204_NO_CONTENT)
                
        except Exception as e:
            print(f"Error in destroy view method: {str(e)}")
            return Response(
                {'error': f'Failed to delete template: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['post'])
    def delete(self, request, pk=None):
        """
        Альтернативный метод удаления шаблона через POST запрос
        """
        return self.destroy(request, pk=pk)

    @action(detail=True, methods=['post'])
    def destroy_template(self, request, pk=None):
        """
        Альтернативный метод удаления шаблона через POST запрос на /destroy_template/
        """
        try:
            instance = self.get_object()
            template_id = instance.id
            template_name = instance.name
            
            # Проверяем права доступа (только создатель или админ может удалить)
            if not request.user.is_superuser and instance.created_by != request.user:
                return Response(
                    {'error': 'You do not have permission to delete this template'}, 
                    status=status.HTTP_403_FORBIDDEN
                )
                
            # Логируем попытку удаления
            print(f"Marking template {template_id} '{template_name}' as deleted by {request.user.username}")
            
            # Вместо физического удаления, помечаем запись как удаленную
            instance.is_deleted = True
            instance.save()
            
            # Логируем успешное удаление
            print(f"Template {template_id} '{template_name}' was successfully marked as deleted")
            
            return Response(
                {"success": True, "message": f"Template {template_name} marked as deleted successfully"}, 
                status=status.HTTP_200_OK
            )
                
        except Exception as e:
            print(f"Error in destroy_template action: {str(e)}")
            return Response(
                {'error': f'Failed to process request: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=True, methods=['post'])
    def generate_pdf(self, request, pk=None):
        template = self.get_object()
        config = template.configuration
        report_data = self.generate_report_data(config)
        return JsonResponse({
            'report_name': template.name,
            'project_name': template.project.name,
            'generated_by': request.user.username,
            'generated_at': timezone.now().isoformat(),
            'report_data': report_data
        })
        
    @action(detail=True, methods=['post'])
    def generate_report(self, request, pk=None):
        """
        Создает отчет на основе шаблона с учетом выбранных параметров
        """
        template = self.get_object()
        project_id = request.data.get('project')
        if not project_id:
            project_id = template.project_id
        
        time_range = request.data.get('time_range', '30d')
        
        # Определяем временной диапазон
        end_date = timezone.now()
        if time_range == '24h':
            start_date = end_date - timedelta(hours=24)
        elif time_range == '7d':
            start_date = end_date - timedelta(days=7)
        elif time_range == '30d':
            start_date = end_date - timedelta(days=30)
        elif time_range == '3m':
            start_date = end_date - timedelta(days=90)
        elif time_range == '6m':
            start_date = end_date - timedelta(days=180)
        elif time_range == '1y':
            start_date = end_date - timedelta(days=365)
        elif time_range == 'custom':
            try:
                start_date = datetime.strptime(request.data.get('start_date'), '%Y-%m-%d')
                end_date = datetime.strptime(request.data.get('end_date'), '%Y-%m-%d')
            except (ValueError, TypeError):
                return Response(
                    {'error': 'Invalid date format for custom range. Use YYYY-MM-DD format.'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
        else:
            return Response(
                {'error': 'Invalid time range. Must be one of: 24h, 7d, 30d, 3m, 6m, 1y, custom'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Генерируем данные отчета
        config = template.configuration.copy() if template.configuration else {}
        config['project_id'] = project_id
        config['start_date'] = start_date.isoformat()
        config['end_date'] = end_date.isoformat()
        
        try:
            # Получаем тест-кейсы и тест-ранны
            project = Project.objects.get(id=project_id)
            test_cases = TestCase.objects.filter(folder__project=project)
            test_runs = TestRun.objects.filter(
                test_case__in=test_cases,
                started_at__range=(start_date, end_date)
            ).order_by('-started_at')
            
            # Общая статистика
            total_tests = test_cases.count()
            total_runs = test_runs.count()
            successful_runs = test_runs.filter(status='success').count()
            success_rate = 0
            if total_runs > 0:
                success_rate = int((successful_runs / total_runs) * 100)
            
            # Средняя длительность выполнения тестов
            avg_duration = 0
            if total_runs > 0:
                duration_sum = sum(run.duration or 0 for run in test_runs if run.duration is not None)
                if duration_sum > 0:
                    avg_duration = duration_sum / total_runs
            
            # Последние тест-раны
            recent_runs = []
            for run in test_runs[:10]:
                recent_runs.append({
                    'test_name': run.test_case.title,
                    'status': run.status,
                    'duration': f"{run.duration:.1f}s" if run.duration else "-",
                    'date': run.started_at.strftime('%d.%m.%Y %H:%M') if run.started_at else "-"
                })
            
            # Данные по трендам выполнения
            execution_data = self.get_execution_trend_data(project_id, start_date, end_date)
            
            # Распределение статусов
            status_data = self.get_status_distribution_data(project_id, start_date, end_date)
            
            return Response({
                'id': template.id,
                'name': template.name,
                'project_name': project.name,
                'generated_by': request.user.username,
                'generated_at': timezone.now().isoformat(),
                'time_range': time_range,
                'start_date': start_date.isoformat(),
                'end_date': end_date.isoformat(),
                'metrics': {
                    'totalTests': total_tests,
                    'successfulTests': f"{success_rate}%",
                    'avgTime': f"{avg_duration:.1f}s"
                },
                'charts': {
                    'execution_trend': execution_data,
                    'status_distribution': status_data
                },
                'recent_test_runs': recent_runs
            })
            
        except Exception as e:
            print(f"Error generating report: {str(e)}", exc_info=True)
            return Response(
                {'error': f'Error generating report: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def generate_report_data(self, config):
        """
        Генерирует данные отчета на основе конфигурации
        """
        project_id = config.get('project_id')
        if not project_id:
            return {}
            
        try:
            # Преобразуем строковые даты в datetime
            start_date = datetime.fromisoformat(config.get('start_date', ''))
            end_date = datetime.fromisoformat(config.get('end_date', ''))
        except (ValueError, TypeError):
            # Если даты не указаны или некорректные, используем последние 30 дней
            end_date = timezone.now()
            start_date = end_date - timedelta(days=30)
            
        try:
            # Получаем базовую статистику
            test_stats = self.get_execution_trend_data(project_id, start_date, end_date)
            status_stats = self.get_status_distribution_data(project_id, start_date, end_date)
            
            # Формируем структуру данных отчета
            report_data = {
                'metrics': self.get_metrics_data(project_id, start_date, end_date),
                'charts': {
                    'execution_trend': test_stats,
                    'status_distribution': status_stats
                },
                'recent_test_runs': self.get_recent_test_runs(project_id, 10)
            }
            
            return report_data
        except Exception as e:
            print(f"Error generating report data: {str(e)}", exc_info=True)
            return {}

    def get_metrics_data(self, project_id, start_date, end_date):
        """
        Получить основные метрики для отчета
        """
        try:
            # Получаем тест-кейсы и тест-ранны
            test_cases = TestCase.objects.filter(folder__project_id=project_id)
            test_runs = TestRun.objects.filter(
                test_case__in=test_cases,
                started_at__range=(start_date, end_date)
            )
            
            # Общая статистика
            total_tests = test_cases.count()
            total_runs = test_runs.count()
            successful_runs = test_runs.filter(status='success').count()
            success_rate = 0
            if total_runs > 0:
                success_rate = int((successful_runs / total_runs) * 100)
            
            # Средняя длительность выполнения тестов
            avg_duration = 0
            if total_runs > 0:
                duration_sum = sum(run.duration or 0 for run in test_runs if run.duration is not None)
                if duration_sum > 0:
                    avg_duration = duration_sum / total_runs
            
            return {
                'totalTests': total_tests,
                'successfulTests': f"{success_rate}%",
                'avgTime': f"{avg_duration:.1f}s"
            }
        except Exception as e:
            print(f"Error getting metrics data: {str(e)}", exc_info=True)
            return {
                'totalTests': 0,
                'successfulTests': "0%",
                'avgTime': "0s"
            }
            
    def get_execution_trend_data(self, project_id, start_date, end_date):
        """
        Получить данные о тренде выполнения тестов
        """
        try:
            # Получаем периоды для анализа
            days = (end_date - start_date).days + 1
            dates = []
            current_date = start_date
            while current_date <= end_date:
                dates.append(current_date.date())
                current_date += timedelta(days=1)
                
            # Получаем тест-ранны
            test_runs = TestRun.objects.filter(
                test_case__folder__project_id=project_id,
                started_at__range=(start_date, end_date)
            )
            
            # Группируем по датам и статусам
            successful = []
            failed = []
            
            for date in dates:
                day_runs = test_runs.filter(started_at__date=date)
                successful.append(day_runs.filter(status='success').count())
                failed.append(day_runs.filter(status__in=['failure', 'error']).count())
                
            # Форматируем даты для отображения
            formatted_dates = [date.strftime('%d.%m') for date in dates]
            
            return {
                'labels': formatted_dates,
                'datasets': [
                    {
                        'label': 'Успешные',
                        'data': successful,
                        'borderColor': '#10B981',
                        'backgroundColor': 'rgba(16, 185, 129, 0.1)'
                    },
                    {
                        'label': 'Неуспешные',
                        'data': failed,
                        'borderColor': '#EF4444',
                        'backgroundColor': 'rgba(239, 68, 68, 0.1)'
                    }
                ]
            }
        except Exception as e:
            print(f"Error getting execution trend data: {str(e)}", exc_info=True)
            return {
                'labels': [],
                'datasets': []
            }
            
    def get_status_distribution_data(self, project_id, start_date, end_date):
        """
        Получить данные о распределении статусов тестов
        """
        try:
            # Получаем тест-ранны
            test_runs = TestRun.objects.filter(
                test_case__folder__project_id=project_id,
                started_at__range=(start_date, end_date)
            )
            
            # Подсчитываем количество для каждого статуса
            success_count = test_runs.filter(status='success').count()
            failure_count = test_runs.filter(status='failure').count()
            error_count = test_runs.filter(status='error').count()
            other_count = test_runs.exclude(status__in=['success', 'failure', 'error']).count()
            
            return {
                'labels': ['Успешно', 'Ошибка', 'Неудача', 'Другие'],
                'datasets': [{
                    'data': [success_count, error_count, failure_count, other_count],
                    'backgroundColor': [
                        '#10B981',  # Зеленый - успешно
                        '#EF4444',  # Красный - ошибка
                        '#F59E0B',  # Оранжевый - неудача
                        '#6B7280'   # Серый - другие
                    ]
                }]
            }
        except Exception as e:
            print(f"Error getting status distribution data: {str(e)}", exc_info=True)
            return {
                'labels': [],
                'datasets': []
            }
            
    def get_recent_test_runs(self, project_id, limit=10):
        """
        Получить последние выполненные тесты
        """
        try:
            # Получаем последние тест-ранны
            test_runs = TestRun.objects.filter(
                test_case__folder__project_id=project_id
            ).order_by('-started_at')[:limit]
            
            # Формируем данные
            recent_runs = []
            for run in test_runs:
                recent_runs.append({
                    'test_name': run.test_case.title if run.test_case else 'Unknown',
                    'status': run.status,
                    'duration': f"{run.duration:.1f}s" if run.duration else "-",
                    'date': run.started_at.strftime('%d.%m.%Y %H:%M') if run.started_at else "-"
                })
                
            return recent_runs
        except Exception as e:
            print(f"Error getting recent test runs: {str(e)}", exc_info=True)
            return []

    @action(detail=True, methods=['post'])
    def destroy_template(self, request, pk=None):
        """
        Альтернативный метод удаления шаблона через POST запрос на /destroy/
        """
        try:
            instance = self.get_object()
            template_id = instance.id
            template_name = instance.name
            
            # Проверяем права доступа
            if not request.user.is_superuser and instance.created_by != request.user:
                return Response(
                    {'error': 'You do not have permission to delete this template'}, 
                    status=status.HTTP_403_FORBIDDEN
                )
            
            print(f"Trying to destroy template {template_id} '{template_name}' via POST method")
            
            # Пробуем напрямую удалить из базы данных
            from django.db import connection
            cursor = connection.cursor()
            
            # Лог SQL-запроса
            query = f"DELETE FROM flowtestapp_reporttemplate WHERE id = {template_id}"
            print(f"Executing direct SQL: {query}")
            
            try:
                cursor.execute(query)
                print(f"SQL delete successful, rows affected: {cursor.rowcount}")
                return Response({"success": True, "message": f"Template {template_name} deleted successfully"}, 
                               status=status.HTTP_200_OK)
            except Exception as sql_error:
                print(f"SQL Error: {str(sql_error)}")
                return Response(
                    {'error': f'Failed to delete template via SQL: {str(sql_error)}'}, 
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )
                
        except Exception as e:
            print(f"Error in destroy_template action: {str(e)}")
            return Response(
                {'error': f'Failed to process request: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

# Аналитика
class AnalyticsView(APIView):
    def get(self, request):
        period = request.query_params.get('period', '24h')
        now = timezone.now()
        if period == '24h':
            start_date = now - timedelta(hours=24)
        elif period == '7d':
            start_date = now - timedelta(days=7)
        elif period == '30d':
            start_date = now - timedelta(days=30)
        else:
            return Response({'error': 'Invalid period'}, status=400)
        test_runs = TestRun.objects.filter(started_at__gte=start_date)
        metrics = {
            'success_rate': test_runs.filter(status='passed').count() / test_runs.count() * 100 if test_runs.exists() else 0,
            'avg_duration': test_runs.aggregate(Avg('duration'))['duration__avg'] or 0,
            'total_tests': test_runs.count()
        }
        trends_data = (
            test_runs.annotate(date=F('started_at__date'))
            .values('date')
            .annotate(success=Count('id', filter=Q(status='passed')), failed=Count('id', filter=Q(status='failed')))
            .order_by('date')
        )
        distribution_data = {
            'passed': test_runs.filter(status='passed').count(),
            'failed': test_runs.filter(status='failed').count(),
            'skipped': test_runs.filter(status='skipped').count()
        }
        charts = {
            'trends': {
                'labels': [item['date'].strftime('%Y-%m-%d') for item in trends_data],
                'success': [item['success'] for item in trends_data],
                'failed': [item['failed'] for item in trends_data]
            },
            'distribution': distribution_data
        }
        results = TestReportSerializer(test_runs, many=True).data
        response_data = {'metrics': metrics, 'charts': charts, 'results': results}
        serializer = AnalyticsResponseSerializer(response_data)
        return Response(serializer.data)

# Экспорт отчётов
class ReportExportView(APIView):
    def get(self, request, format):
        period = request.query_params.get('period', '24h')
        now = timezone.now()
        if period == '24h':
            start_date = now - timedelta(hours=24)
        elif period == '7d':
            start_date = now - timedelta(days=7)
        elif period == '30d':
            start_date = now - timedelta(days=30)
        else:
            return Response({'error': 'Invalid period'}, status=400)
        test_runs = TestRun.objects.filter(started_at__gte=start_date)
        if format == 'excel':
            return self.export_excel(test_runs)
        elif format == 'pdf':
            return self.export_pdf(test_runs)
        return Response({'error': 'Invalid format'}, status=400)

    def export_excel(self, test_runs):
        import pandas as pd
        df = pd.DataFrame(list(test_runs.values('test_case', 'status', 'duration', 'started_at', 'finished_at')))
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            df.to_excel(writer, sheet_name='Test Results', index=False)
            workbook = writer.book
            worksheet = writer.sheets['Test Results']
            header_font = Font(bold=True,
                text_wrap=True,
                valign='top',
                bg_color='#D9EAD3',
                border=1
            )
            for col_num, value in enumerate(df.columns.values):
                worksheet.write(0, col_num, value, header_font)
                worksheet.set_column(col_num, col_num, 15)
        output.seek(0)
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=test_report_{timezone.now().strftime("%Y%m%d")}.xlsx'
        return response

    def export_pdf(self, test_runs):
        from django.template.loader import get_template
        import pdfkit
        template = get_template('report_template.html')
        context = {
            'test_runs': test_runs,
            'generated_at': timezone.now(),
            'total_tests': test_runs.count(),
            'passed_tests': test_runs.filter(status='passed').count(),
            'failed_tests': test_runs.filter(status='failed').count()
        }
        html = template.render(context)
        pdf = pdfkit.from_string(html, False)
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename=test_report_{timezone.now().strftime("%Y%m%d")}.pdf'
        return response

# Детали отчёта
class ReportDetailView(APIView):
    def get(self, request, report_id):
        try:
            test_run = TestRun.objects.get(id=report_id)
            serializer = TestReportSerializer(test_run)
            return Response(serializer.data)
        except TestRun.DoesNotExist:
            return Response({'error': 'Report not found'}, status=status.HTTP_404_NOT_FOUND)

# Папки и тест-кейсы проекта
class ProjectFoldersAndTestCases(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, pk):
        try:
            project = Project.objects.get(pk=pk)
            folders = Folder.objects.filter(project=project)
            test_cases = TestCase.objects.filter(project=project)
            folder_serializer = FolderSerializer(folders, many=True)
            test_case_serializer = TestCaseSerializer(test_cases, many=True)
            return Response({'folders': folder_serializer.data, 'test_cases': test_case_serializer.data})
        except Project.DoesNotExist:
            return Response({'error': 'Project not found'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# Получение профиля пользователя
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_user_profile(request):
    try:
        user = request.user
        serializer = CustomUserSerializer(user, context={'request': request})
        return Response(serializer.data)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# Создание пользователя
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def create_user(request):
    """
    Create a new user (admin only)
    """
    # Only staff/admin users can create new users
    if not (request.user.is_staff or request.user.is_superuser):
        return Response({"detail": "You do not have permission to perform this action."}, 
                     status=status.HTTP_403_FORBIDDEN)
    
    try:
        serializer = CustomUserSerializer(data=request.data, context={'request': request})
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# Загрузка аватара
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def upload_avatar(request):
    if not request.FILES or 'file' not in request.FILES:
        return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)
    file = request.FILES['file']
    if file.size > 2 * 1024 * 1024:
        return Response({'error': 'File too large'}, status=status.HTTP_400_BAD_REQUEST)
    allowed_types = ['image/jpeg', 'image/png', 'image/gif']
    if file.content_type not in allowed_types:
        return Response({'error': 'Invalid file type'}, status=status.HTTP_400_BAD_REQUEST)
    try:
        user = request.user
        if user.avatar:
            if user.avatar.storage.exists(user.avatar.name):
                user.avatar.delete(save=False)
        user.avatar = file  # Используем upload_to из модели
        user.save()
        serializer = CustomUserSerializer(user, context={'request': request})
        return Response({
            'success': True,
            'message': 'Avatar updated successfully',
            'user': serializer.data
        })
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# Статистика создания тест-кейсов
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def test_cases_creation_stats(request):
    try:
        days = int(request.GET.get('days', 30))
        project_id = request.GET.get('project_id')
        
        queryset = TestCase.objects.all()
        if project_id:
            queryset = queryset.filter(folder__project_id=project_id)
            
        end_date = timezone.now()
        start_date = end_date - timedelta(days=days)
        
        # Получаем все даты в диапазоне
        dates = []
        current_date = start_date.date()
        while current_date <= end_date.date():
            dates.append(current_date)
            current_date += timedelta(days=1)
            
        # Получаем статистику создания тест-кейсов
        stats = queryset.filter(
            created_at__range=(start_date, end_date)
        ).annotate(
            date=TruncDate('created_at')
        ).values('date').annotate(
            count=Count('id')
        ).order_by('date')
        
        # Преобразуем в словарь для быстрого доступа
        stats_dict = {stat['date']: stat['count'] for stat in stats}
        
        # Формируем финальный результат
        result = []
        for date in dates:
            result.append({
                'date': date.isoformat(),
                'count': stats_dict.get(date, 0)
            })
        
        return Response(result)
    except Exception as e:
        return Response({'error': str(e)}, status=400)

# Исправляем tests_over_time
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def tests_over_time(request):
    try:
        project_id = request.query_params.get('project_id')
        if not project_id:
            return Response({'error': 'project_id is required'}, status=400)

        # Получаем период для отображения (по умолчанию 30 дней)
        days = int(request.query_params.get('days', 30))
        end_date = timezone.now()
        start_date = end_date - timedelta(days=days)

        # Получаем все даты в диапазоне
        dates = []
        current_date = start_date.date()
        while current_date <= end_date.date():
            dates.append(current_date)
            current_date += timedelta(days=1)

        # Get test runs for the project
        test_runs = TestRun.objects.filter(
            test_case__folder__project_id=project_id,
            started_at__range=(start_date, end_date)
        )

        # Group test runs by date and calculate statistics
        daily_stats = test_runs.annotate(
            date=TruncDate('started_at')
        ).values('date').annotate(
            total=Count('id'),
            passed=Count('id', filter=Q(status='passed')),
            failed=Count('id', filter=Q(status='failed')),
            error=Count('id', filter=Q(status='error'))
        ).order_by('date')

        # Преобразуем в словарь для быстрого доступа
        stats_dict = {stat['date']: stat for stat in daily_stats}
        
        # Формируем данные для всех дат в диапазоне
        formatted_dates = []
        success_rates = []
        
        for date in dates:
            formatted_dates.append(date.strftime('%d.%m.%Y'))
            
            if date in stats_dict:
                stat = stats_dict[date]
                if stat['total'] > 0:
                    success_rate = (stat['passed'] / stat['total']) * 100
                else:
                    success_rate = 0
            else:
                success_rate = 0
                
            success_rates.append(int(success_rate))

        # Результат в формате для фронтенда
        response_data = {
            'labels': formatted_dates,
            'success_rate': success_rates
        }

        return Response(response_data)
    except Exception as e:
        print(f"Error in tests_over_time: {str(e)}")
        return Response({'labels': [], 'success_rate': []})

# Распределение результатов
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def results_distribution(request):
    try:
        project_id = request.query_params.get('project_id')
        if not project_id:
            return Response([])

        # Получаем все тест-раны для данного проекта
        test_runs = TestRun.objects.filter(test_case__folder__project_id=project_id)
        
        # Группируем и считаем по статусам
        status_counts = test_runs.values('status').annotate(count=Count('id'))
        
        # Преобразуем в нужный формат для фронтенда
        result = []
        for item in status_counts:
            result.append({
                'status': item['status'],
                'count': item['count']
            })
            
        return Response(result)
    except Exception as e:
        print(f"Error in results_distribution: {str(e)}")
        return Response([])

# Распределение по приоритетам
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def priority_distribution(request):
    try:
        project_id = request.query_params.get('project_id')
        if not project_id:
            return Response([])

        # Получаем все тест-кейсы для данного проекта
        test_cases = TestCase.objects.filter(folder__project_id=project_id)
        
        # Группируем и считаем по приоритетам
        priority_counts = test_cases.values('priority').annotate(count=Count('id'))
        
        # Преобразуем в нужный формат для фронтенда
        result = []
        for item in priority_counts:
            result.append({
                'priority': item['priority'],
                'count': item['count']
            })
            
        return Response(result)
    except Exception as e:
        print(f"Error in priority_distribution: {str(e)}")
        return Response([])

# Нестабильность тестов
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def test_flakiness(request):
    try:
        project_id = request.query_params.get('project_id')
        if not project_id:
            return Response([])

        # Get test cases for the project
        queryset = TestCase.objects.filter(folder__project_id=project_id)

        # Find tests with alternating success/failure patterns
        flaky_tests = []
        for test in queryset:
            runs = TestRun.objects.filter(test_case=test).order_by('-started_at')[:10]
            if runs.count() >= 5:  # Minimum 5 runs for analysis
                statuses = list(runs.values_list('status', flat=True))
                changes = sum(1 for i in range(len(statuses)-1) if statuses[i] != statuses[i+1])
                if changes >= 2:  # Minimum 2 status changes
                    flaky_tests.append({
                        'test_id': test.id,
                        'title': test.title,
                        'changes': changes,
                        'last_runs': statuses
                    })

        return Response(sorted(flaky_tests, key=lambda x: x['changes'], reverse=True))
    except Exception as e:
        print(f"Error in test_flakiness: {str(e)}")
        return Response([])

# Статистика выполнения тестов
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def test_execution_stats(request, project_id=None):
    try:
        if not project_id:
            project_id = request.query_params.get('project_id')
            if not project_id:
                return Response({
                    'total_tests': 0,
                    'total_executions': 0,
                    'status_distribution': {
                        'passed': 0,
                        'failed': 0,
                        'error': 0,
                        'running': 0,
                        'pending': 0
                    }
                })

        # Count test cases
        total_tests = TestCase.objects.filter(folder__project_id=project_id).count()

        # Get all test runs
        test_runs = TestRun.objects.filter(test_case__folder__project_id=project_id)
        total_executions = test_runs.count()

        # Calculate distribution by status
        status_distribution = {}
        status_counts = test_runs.values('status').annotate(count=Count('id'))
        for item in status_counts:
            status_distribution[item['status']] = item['count']

        # Ensure all statuses have a value
        for status in ['passed', 'failed', 'error', 'skipped', 'running', 'pending']:
            if status not in status_distribution:
                status_distribution[status] = 0

        # Calculate average execution time (in seconds)
        avg_time = test_runs.aggregate(avg_time=Avg(
            ExpressionWrapper(
                F('finished_at') - F('started_at'),
                output_field=DurationField()
            )
        ))['avg_time']

        avg_execution_time = 0
        if avg_time:
            avg_execution_time = avg_time.total_seconds()

        stats = {
            'total_tests': total_tests,
            'total_executions': total_executions,
            'status_distribution': status_distribution,
            'avg_execution_time': avg_execution_time
        }

        return Response(stats)
    except Exception as e:
        print(f"Error in test_execution_stats: {str(e)}")
        return Response({
            'total_tests': 0,
            'total_executions': 0,
            'status_distribution': {
                'passed': 0,
                'failed': 0,
                'error': 0,
                'running': 0,
                'pending': 0
            },
            'avg_execution_time': 0
        })